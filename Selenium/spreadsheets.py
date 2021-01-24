from openpyxl import load_workbook


class SheetHandler:
    def __init__(self):
        self.workbook = load_workbook(filename=self.FILENAME)
        self.sheet = self.workbook.active

    def save(self):
        self.workbook.save(filename=self.FILENAME)


class EmployeeDataSheetHandler(SheetHandler):
    FILENAME = "EmployeeData100.xlsx"
    MIN_ROW = 2
    MAX_COL = 9

    def iterate_employees(self):
        for row in self.sheet.iter_rows(min_row=self.MIN_ROW, max_col=self.MAX_COL):
            yield row

    def write_cell(self, cell, value):
        cell.set_explicit_value(value)


class InvoiceSheetHandler(SheetHandler):
    FILENAME = "InvoicesData.xlsx"

    def write_row(self, invoice_data: dict):
        data = [
            invoice_data.get('invoice_id'),
            invoice_data.get('date'),
            invoice_data.get('customer_address'),
            invoice_data.get('customer_name'),
            invoice_data.get('customer_phone'),
            invoice_data.get('amount'),
        ]
        self.sheet.append(data)


class StoreDataSheetHandler(SheetHandler):
    FILENAME = "DrugStores.xlsx"

    def choose_sheet(self, store: str):
        try:
            self.sheet = self.workbook[store]
        except KeyError:
            self.sheet = self.workbook.active

    def write_data(self, scraped_data: list):
        print(scraped_data[:10])
        for item in scraped_data:
            self.sheet.append(item)
